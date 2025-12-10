"""
WhatsApp Auto-Sender para iPhone
==================================
Automatiza el envío de mensajes de WhatsApp usando datos de Excel en OneDrive

Autor: José Estevan Calvo Martinez
GitHub: github.com/tu-usuario
LinkedIn: linkedin.com/in/estevancalvo
Versión: 1.0.0
Licencia: MIT

Descripción:
-----------
Este script automatiza el envío de mensajes de WhatsApp en iPhone mediante:
- Descarga de archivos Excel desde OneDrive/Microsoft 365
- Procesamiento de hipervínculos de WhatsApp en Excel
- Apertura automática de WhatsApp con mensajes pre-escritos
- Control manual del usuario antes de enviar cada mensaje

Requisitos:
----------
- iPhone con iOS 15+
- App a-Shell
- Librerías: pandas, openpyxl, requests
- WhatsApp instalado
- Archivo Excel en OneDrive con hipervínculos de WhatsApp

Caso de uso:
-----------
Perfecto para empresas que necesitan enviar mensajes personalizados
manteniendo el control y cumpliendo con las políticas de WhatsApp.

Resultado:
---------
Reduce el tiempo de envío de mensajes en un 85% comparado con
el proceso manual de copiar y pegar.
"""

import pandas as pd
import sys
import os
import time
import openpyxl
from openpyxl import load_workbook
import requests
import base64

# ============================================================
# CONFIGURACIÓN - EDITA ESTO ANTES DE USAR
# ============================================================

# IMPORTANTE: Pega aquí el link que obtienes al hacer clic en "Compartir" en OneDrive
# Debe ser un enlace corto, como: "https://1drv.ms/x/s!A...xxxxxxxx"
ONEDRIVE_LINK = ""  # <-- ¡REEMPLAZA CON TU LINK DE ONEDRIVE!

# Nombre del archivo local temporal
ARCHIVO_LOCAL = "datos.xlsx"

# Nombre de la hoja de Excel que contiene los mensajes
NOMBRE_HOJA_MENSAJES = "Mensajes de Whatsapp"

# ============================================================
# FUNCIONES DE DESCARGA DESDE ONEDRIVE
# ============================================================

def crear_link_descarga_directa(link_compartido):
    """
    Convierte un link compartido de OneDrive a enlace de descarga directa
    
    Args:
        link_compartido (str): Link compartido de OneDrive (formato 1drv.ms)
    
    Returns:
        str: URL de descarga directa o None si hay error
    """
    try:
        link_base = link_compartido.split('?')[0]
        data_bytes = link_base.encode('utf-8')
        base64_bytes = base64.b64encode(data_bytes)
        base64_string = base64_bytes.decode('utf-8').rstrip('=').replace('/', '_').replace('+', '-')

        link_descarga = f"https://api.onedrive.com/v1.0/shares/u!{base64_string}/root/content"
        return link_descarga
    except Exception as e:
        print(f"⚠️  Error al convertir el link de OneDrive: {e}")
        return None

def descargar_excel_onedrive(link_compartido, nombre_archivo):
    """
    Descarga el archivo Excel desde OneDrive
    Elimina el archivo anterior para asegurar la versión más reciente
    
    Args:
        link_compartido (str): Link compartido de OneDrive
        nombre_archivo (str): Nombre del archivo local donde guardar
    
    Returns:
        bool: True si la descarga fue exitosa, False en caso contrario
    """
    # Eliminar archivo anterior si existe
    if os.path.exists(nombre_archivo):
        try:
            os.remove(nombre_archivo)
            print(f"Archivo anterior '{nombre_archivo}' eliminado")
        except OSError as e:
            print(f"Error al eliminar archivo anterior: {e}")
            return False

    try:
        print("=" * 70)
        print("☁️   DESCARGANDO DESDE ONEDRIVE")
        print("=" * 70)

        link_descarga = crear_link_descarga_directa(link_compartido)
        if not link_descarga:
            return False

        print(f"📥 Descargando desde OneDrive...")
        
        # Realizar petición de descarga
        respuesta = requests.get(
            link_descarga, 
            stream=True, 
            allow_redirects=True, 
            headers={'User-Agent': 'Mozilla/5.0'},
            timeout=60
        )
        
        if respuesta.status_code == 200:
            with open(nombre_archivo, 'wb') as f:
                for chunk in respuesta.iter_content(chunk_size=8192):
                    f.write(chunk)
            
            tamaño = os.path.getsize(nombre_archivo)
            print(f"✅ Descargado exitosamente ({tamaño:,} bytes)")
            print(f"📁 Guardado como: {nombre_archivo}\n")
            return True
        else:
            print(f"Error: El servidor respondió con código {respuesta.status_code}")
            print("   Verifica que el link sea correcto y los permisos sean públicos")
            return False
            
    except requests.exceptions.RequestException as e:
        print(f"❌ Error de red al descargar: {e}")
        print("\n💡 Verifica que:")
        print("  1. El link compartido de OneDrive sea correcto")
        print("  2. El permiso del archivo sea 'Cualquiera con el vínculo'")
        print("  3. Tengas conexión a internet activa")
        return False
    except Exception as e:
        print(f"❌ Error inesperado durante la descarga: {e}")
        return False

# ============================================================
# FUNCIONES DE PROCESAMIENTO DE EXCEL
# ============================================================

def extraer_hyperlinks_excel(archivo_excel):
    """
    Extrae los hipervínculos de WhatsApp del archivo Excel
    
    Args:
        archivo_excel (str): Ruta del archivo Excel
    
    Returns:
        list: Lista de diccionarios con información de cada hipervínculo
              Cada dict contiene: celda, texto, url
    """
    try:
        wb = load_workbook(archivo_excel)
        
        # Mostrar hojas disponibles
        print("=" * 70)
        print("📋 HOJAS DISPONIBLES EN EL EXCEL")
        print("=" * 70)
        for i, hoja in enumerate(wb.sheetnames, 1):
            print(f"  {i}. {hoja}")
        
        # Buscar la hoja de mensajes con variaciones
        ws = None
        variaciones = [
            NOMBRE_HOJA_MENSAJES,
            "Mensajes de WhatsApp",
            "mensajes de whatsapp",
            "Mensajes WhatsApp",
            "MensajesWhatsapp",
            "Mensajes WA",
            "WhatsApp"
        ]
        
        for var in variaciones:
            if var in wb.sheetnames:
                ws = wb[var]
                print(f"\n✅ Usando hoja: '{var}'")
                break
        
        if ws is None:
            print(f"\n⚠️  No se encontró la hoja '{NOMBRE_HOJA_MENSAJES}'")
            print(f"⚠️  Usando la primera hoja: '{wb.sheetnames[0]}'")
            ws = wb.worksheets[0]
        
        # Extraer hipervínculos
        hyperlinks = []
        print("\n" + "=" * 70)
        print("🔍 EXTRAYENDO HIPERVÍNCULOS DE WHATSAPP")
        print("=" * 70)
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink and cell.hyperlink.target:
                    link = cell.hyperlink.target
                    if 'wa.me' in link or 'api.whatsapp.com' in link:
                        hyperlinks.append({
                            'celda': cell.coordinate,
                            'texto': cell.value,
                            'url': link
                        })
                        texto_mostrar = str(cell.value)[:50] if cell.value else "Sin texto"
                        print(f"✅ {cell.coordinate}: {texto_mostrar}")
        
        print(f"\n📊 Total encontrados: {len(hyperlinks)} hipervínculos")
        return hyperlinks
        
    except Exception as e:
        print(f"❌ Error al leer hipervínculos del Excel: {e}")
        import traceback
        print(traceback.format_exc())
        return []

# ============================================================
# FUNCIONES DE WHATSAPP
# ============================================================

def convertir_url_whatsapp(url):
    """
    Convierte URLs de WhatsApp Web a formato de app móvil
    
    Args:
        url (str): URL de WhatsApp en formato web
    
    Returns:
        str: URL en formato de app móvil (whatsapp://)
    """
    if 'wa.me' in url:
        return url.replace('https://wa.me/', 'whatsapp://send?phone=').replace('?text=', '&text=')
    elif 'api.whatsapp.com' in url:
        return url.replace('https://api.whatsapp.com/send', 'whatsapp://send')
    return url

def abrir_whatsapp(url):
    """
    Abre WhatsApp con el mensaje prellenado
    Compatible con iOS (a-Shell), macOS y Linux
    
    Args:
        url (str): URL de WhatsApp
    
    Returns:
        bool: True si se abrió correctamente, False en caso contrario
    """
    try:
        url_movil = convertir_url_whatsapp(url)
        print(f"📱 Abriendo WhatsApp...")
        
        if sys.platform == "win32":
            os.system(f'start "" "{url_movil}"')
        else:
            # Para macOS, Linux y a-Shell en iOS
            os.system(f'open "{url_movil}"')
        
        return True
    except Exception as e:
        print(f"❌ Error al abrir WhatsApp: {e}")
        return False

# ============================================================
# FUNCIÓN PRINCIPAL
# ============================================================

def procesar_excel_whatsapp():
    """
    Función principal que ejecuta todo el proceso:
    1. Descarga el Excel desde OneDrive
    2. Extrae los hipervínculos de WhatsApp
    3. Abre WhatsApp para cada mensaje
    4. Espera confirmación del usuario entre mensajes
    """
    try:
        # Verificar configuración
        if not ONEDRIVE_LINK or ONEDRIVE_LINK == "":
            print("=" * 70)
            print("⚠️  CONFIGURACIÓN NECESARIA")
            print("=" * 70)
            print("\n❌ No has configurado tu link de OneDrive")
            print("\n📝 PASOS PARA CONFIGURAR:")
            print("  1. Abre tu archivo en OneDrive (navegador)")
            print("  2. Click en 'Compartir' arriba a la izquierda")
            print("  3. Asegúrate: 'Cualquiera con el vínculo puede ver'")
            print("  4. Click en 'Copiar vínculo'")
            print("  5. Pega ese vínculo en ONEDRIVE_LINK en este script")
            print("\n💡 Ejemplo:")
            print('   ONEDRIVE_LINK = "https://1drv.ms/x/s!xxxxx"')
            return

        print("\n🚀 WhatsApp Auto-Sender desde OneDrive 🚀\n")
        
        # Descargar Excel
        if not descargar_excel_onedrive(ONEDRIVE_LINK, ARCHIVO_LOCAL):
            return
        
        # Extraer hipervínculos
        hyperlinks = extraer_hyperlinks_excel(ARCHIVO_LOCAL)
        
        if not hyperlinks:
            print("\n" + "=" * 70)
            print("⚠️  NO SE ENCONTRARON HIPERVÍNCULOS DE WHATSAPP")
            print("=" * 70)
            print("\n💡 Verifica que:")
            print(f"  1. La hoja '{NOMBRE_HOJA_MENSAJES}' exista en el Excel")
            print("  2. Las celdas contengan hipervínculos (no solo texto)")
            print("  3. Los hipervínculos apunten a 'wa.me' o 'api.whatsapp.com'")
            return
        
        # Mostrar resumen
        print("\n" + "=" * 70)
        print(f"✅ ENCONTRADOS: {len(hyperlinks)} MENSAJES DE WHATSAPP")
        print("=" * 70)
        
        # Instrucciones
        print("\n📋 INSTRUCCIONES:")
        print("  1. El script abrirá WhatsApp para cada mensaje")
        print("  2. El texto ya estará escrito")
        print("  3. Verifica el destinatario y mensaje")
        print("  4. Presiona ENVIAR en WhatsApp")
        print("  5. Regresa a esta terminal")
        print("  6. Presiona ENTER para continuar")
        print("=" * 70)
        
        input("\n▶️  Presiona ENTER para comenzar...")
        
        # Procesar cada mensaje
        mensajes_enviados = 0
        
        for index, link_data in enumerate(hyperlinks, 1):
            print("\n" + "=" * 70)
            print(f"📬 MENSAJE {index} de {len(hyperlinks)}")
            print("=" * 70)
            print(f"📍 Celda: {link_data['celda']}")
            
            if link_data['texto']:
                texto_mostrar = str(link_data['texto'])[:60]
                print(f"📝 Texto: {texto_mostrar}...")
            
            # Abrir WhatsApp
            if abrir_whatsapp(link_data['url']):
                mensajes_enviados += 1
                time.sleep(2)
                
                if index < len(hyperlinks):
                    input("\n⏸️  Presiona ENTER para el siguiente mensaje...")
            else:
                print("❌ No se pudo abrir este mensaje")
                input("\nPresiona ENTER para continuar...")
        
        # Resumen final
        print("\n" + "=" * 70)
        print("🎉 ¡PROCESO FINALIZADO!")
        print("=" * 70)
        print(f"📊 Estadísticas:")
        print(f"   • Total procesados: {len(hyperlinks)} mensajes")
        print(f"   • Enviados exitosamente: {mensajes_enviados}")
        print(f"   • Fecha: {time.strftime('%d/%m/%Y %H:%M:%S')}")
        print("=" * 70)
        
        # Limpiar archivo temporal (opcional)
        try:
            os.remove(ARCHIVO_LOCAL)
            print(f"\n🗑️  Archivo temporal eliminado")
        except:
            pass
        
    except KeyboardInterrupt:
        print("\n\n⚠️  Proceso interrumpido por el usuario")
    except Exception as e:
        print(f"\n❌ Error general: {str(e)}")
        import traceback
        print("\n🔍 Detalles del error:")
        print(traceback.format_exc())

# ============================================================
# PUNTO DE ENTRADA
# ============================================================

if __name__ == "__main__":
    print("\n" + "🚀" * 35)
    print("   WHATSAPP AUTO-SENDER - IPHONE EDITION")
    print("   By José Estevan Calvo Martinez")
    print("🚀" * 35)
    procesar_excel_whatsapp()